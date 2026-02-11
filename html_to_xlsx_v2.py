#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Обработчик HTML таблиц результатов соревнований.
Перетащите HTML файлы на exe (Windows) или на .app (macOS) для обработки.
"""

import sys
import os
import re
import platform
import subprocess
import traceback
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict


# --- Определяем режим работы ---
IS_MACOS = platform.system() == 'Darwin'
IS_WINDOWED = not sys.stdout or (hasattr(sys, 'frozen') and IS_MACOS)

# Файл лога для отладки (на рабочем столе)
if IS_MACOS:
    LOG_PATH = os.path.expanduser("~/Desktop/html_to_xlsx_log.txt")
else:
    LOG_PATH = None


def log(message):
    """Пишет в лог-файл и в stdout (если доступен)."""
    try:
        print(message)
    except Exception:
        pass
    if LOG_PATH:
        try:
            with open(LOG_PATH, 'a', encoding='utf-8') as f:
                f.write(f"{datetime.now().strftime('%H:%M:%S')} {message}\n")
        except Exception:
            pass


def notify(title, message):
    """Показать macOS уведомление или вывести в консоль."""
    if IS_MACOS:
        try:
            # Экранируем кавычки
            safe_msg = message.replace('"', '\\"').replace("'", "\\'")
            safe_title = title.replace('"', '\\"').replace("'", "\\'")
            subprocess.run([
                'osascript', '-e',
                f'display notification "{safe_msg}" with title "{safe_title}"'
            ], timeout=5)
        except Exception:
            pass
    log(f"[{title}] {message}")


def notify_error(message):
    """Показать уведомление об ошибке."""
    notify("html_to_xlsx — Ошибка", message)


def wait_before_exit():
    """
    В консольном запуске ждём Enter, чтобы окно не закрывалось сразу.
    В macOS .app (PyInstaller --windowed) stdin отсутствует -> input() вызывает EOFError,
    поэтому просто выходим без зависания.
    """
    if IS_WINDOWED:
        return
    try:
        input("\nНажмите Enter для выхода...")
    except EOFError:
        pass


def get_output_folder(filepaths):
    """Определяет папку для сохранения результатов.
    
    Пытается создать папку рядом с исходными файлами.
    Если не получается (нет прав) — создаёт на Рабочем столе.
    """
    timestamp = datetime.now().strftime("%d.%m.%y %H-%M")
    folder_name = f"общая таблица {timestamp}"
    
    # Попытка 1: рядом с исходными файлами
    first_file_dir = os.path.dirname(os.path.abspath(filepaths[0]))
    output_folder = os.path.join(first_file_dir, folder_name)
    try:
        os.makedirs(output_folder, exist_ok=True)
        # Проверяем что можно писать
        test_file = os.path.join(output_folder, ".test_write")
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        return output_folder, timestamp
    except (OSError, PermissionError):
        log(f"Нет прав записи в {first_file_dir}, пробуем Рабочий стол")
    
    # Попытка 2: Рабочий стол
    if IS_MACOS:
        desktop = os.path.expanduser("~/Desktop")
    else:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    
    output_folder = os.path.join(desktop, folder_name)
    try:
        os.makedirs(output_folder, exist_ok=True)
        return output_folder, timestamp
    except (OSError, PermissionError):
        log(f"Нет прав записи на Рабочий стол")
    
    # Попытка 3: домашняя папка
    output_folder = os.path.join(os.path.expanduser("~"), folder_name)
    os.makedirs(output_folder, exist_ok=True)
    return output_folder, timestamp


def parse_html_file(filepath):
    """Парсит HTML файл и возвращает список (команда, место)."""
    results = []
    
    # Читаем файл как байты
    with open(filepath, 'rb') as f:
        raw_data = f.read()
    
    # Пытаемся определить кодировку из meta-тега
    # Ищем charset в первых 1000 байтах
    header = raw_data[:1000].decode('ascii', errors='ignore').lower()
    
    encoding = None
    if 'charset=windows-1251' in header or 'charset=cp1251' in header:
        encoding = 'cp1251'
    elif 'charset=utf-8' in header:
        encoding = 'utf-8'
    elif 'charset=koi8-r' in header:
        encoding = 'koi8-r'
    
    # Пробуем разные кодировки
    encodings_to_try = []
    if encoding:
        encodings_to_try.append(encoding)
    encodings_to_try.extend(['cp1251', 'utf-8', 'koi8-r', 'latin-1'])
    
    content = None
    for enc in encodings_to_try:
        try:
            content = raw_data.decode(enc)
            # Проверяем, что декодирование дало нормальный результат
            # (нет массовых символов замены)
            if content.count('\ufffd') < 10:
                break
        except (UnicodeDecodeError, LookupError):
            continue
    
    if content is None:
        log(f"Ошибка: не удалось прочитать файл {filepath}")
        return results
    
    soup = BeautifulSoup(content, 'html.parser')
    
    # Ищем таблицы с результатами (те, у которых есть заголовок с bgcolor=silver)
    tables = soup.find_all('table')
    
    for table in tables:
        # Проверяем, что это таблица с результатами (есть заголовок)
        header_row = table.find('tr', bgcolor='silver')
        if not header_row:
            continue
        
        # Берём только первую строку с данными (остальные - дубли)
        rows = table.find_all('tr')
        data_row = None
        for row in rows:
            cells = row.find_all('td')
            if cells and len(cells) >= 10:
                first_cell = cells[0].get_text(strip=True)
                if first_cell == '1':  # Первая строка начинается с 1
                    data_row = row
                    break
        
        if not data_row:
            continue
        
        cells = data_row.find_all('td')
        
        # Определяем шаг (количество ячеек на участника)
        # Ищем где начинается второй участник (ячейка со значением "2")
        step = 10  # по умолчанию
        for i in range(8, min(15, len(cells))):
            cell_text = cells[i].get_text(strip=True)
            if cell_text == '2':
                step = i
                break
        
        # Обрабатываем каждый блок как отдельного участника
        i = 0
        while i < len(cells):
            remaining_cells = len(cells) - i
            
            # Минимум нужно 4 ячейки (№п/п, номер, ФИО, команда)
            if remaining_cells < 4:
                break
            
            # Размер текущего блока (последний может быть короче)
            current_block_size = min(step, remaining_cells)
            participant_cells = cells[i:i+current_block_size]
            
            # Проверяем что первая ячейка - число (порядковый номер)
            first = participant_cells[0].get_text(strip=True)
            if not first.isdigit():
                i += step
                continue
            
            # Индекс 3 - команда/организация
            team = participant_cells[3].get_text(strip=True)
            
            # Место - ищем с конца блока первое значение, которое является числом (местом)
            place = None
            for j in range(len(participant_cells) - 1, 3, -1):
                cell_text = participant_cells[j].get_text(strip=True)
                # Место должно быть числом или специальным значением (н/ф, в/к, дск)
                if cell_text and ':' not in cell_text:
                    # Пропускаем годы рождения (19xx, 20xx)
                    if cell_text.isdigit() and len(cell_text) == 4:
                        year = int(cell_text)
                        if 1900 <= year <= 2100:
                            continue  # это год рождения, пропускаем
                    
                    # Проверяем что это число или спец.значение
                    if cell_text.isdigit():
                        place = cell_text
                        break
                    # Спец. значения: н/ф, в/к, дск, снят — помечаем как "Сошел"
                    cell_lower = cell_text.lower()
                    if any(x in cell_lower for x in ['н/ф', 'в/к', 'дск', 'снят', 'снт', 'дисквал']):
                        place = 'Сошел'
                        break
                    # Если это битая кодировка (пїЅ) — скорее всего статус, помечаем как "Сошел"
                    if 'пїЅ' in cell_text:
                        place = 'Сошел'
                        break
            
            # Если место не найдено, но команда есть — участник сошел
            if place is None and team:
                place = 'Сошел'
            
            if team and place:
                results.append((team, place))
            
            i += step
    
    return results


def extract_sort_key(team_name):
    """Извлекает ключ сортировки из названия команды.
    
    Если есть нумерация (например "1. МТС"), возвращает (число, название).
    Иначе возвращает (бесконечность, название) для алфавитной сортировки.
    """
    # Ищем нумерацию в начале строки
    match = re.match(r'^(\d+)\.\s*(.+)$', team_name)
    if match:
        return (int(match.group(1)), match.group(2).lower())
    return (float('inf'), team_name.lower())


def process_files(filepaths):
    """Обрабатывает все файлы и группирует результаты по командам."""
    teams_data = defaultdict(list)
    
    for filepath in filepaths:
        if not os.path.exists(filepath):
            log(f"Файл не найден: {filepath}")
            continue
        
        log(f"Обработка: {os.path.basename(filepath)}")
        results = parse_html_file(filepath)
        
        for team, place in results:
            teams_data[team].append(place)
    
    return teams_data


def create_xlsx(teams_data, output_path):
    """Создаёт xlsx файл с результатами."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = ["Команда", "Кол-во участников"] + [str(i) for i in range(1, 21)]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Сортировка команд
    sorted_teams = sorted(teams_data.keys(), key=extract_sort_key)
    
    # Заполнение данных
    for row_idx, team in enumerate(sorted_teams, 2):
        places = teams_data[team]
        
        # Команда
        cell = ws.cell(row=row_idx, column=1, value=team)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Количество участников
        cell = ws.cell(row=row_idx, column=2, value=len(places))
        cell.border = thin_border
        cell.alignment = cell_alignment
        
        # Места (максимум 20)
        for place_idx, place in enumerate(places[:20]):
            cell = ws.cell(row=row_idx, column=3 + place_idx, value=place)
            cell.border = thin_border
            cell.alignment = cell_alignment
        
        # Пустые ячейки для оставшихся мест
        for empty_idx in range(len(places), 20):
            cell = ws.cell(row=row_idx, column=3 + empty_idx, value="")
            cell.border = thin_border
    
    # Ширина столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    # Столбцы для мест (C-V = 20 столбцов)
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 
                       'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']:
        ws.column_dimensions[col_letter].width = 6
    
    wb.save(output_path)
    log(f"Файл сохранён: {output_path}")


def open_folder(path):
    """Открывает папку в Finder/Проводнике."""
    try:
        if IS_MACOS:
            subprocess.run(['open', path], timeout=5)
        elif platform.system() == 'Windows':
            os.startfile(path)
    except Exception:
        pass


def main():
    try:
        # Очищаем лог при каждом запуске
        if LOG_PATH:
            try:
                with open(LOG_PATH, 'w', encoding='utf-8') as f:
                    f.write(f"=== html_to_xlsx запуск {datetime.now()} ===\n")
                    f.write(f"sys.argv: {sys.argv}\n")
                    f.write(f"platform: {platform.system()} {platform.machine()}\n")
                    f.write(f"IS_WINDOWED: {IS_WINDOWED}\n\n")
            except Exception:
                pass
        
        if len(sys.argv) < 2:
            notify("html_to_xlsx", "Перетащите HTML файлы на иконку приложения")
            log("Нет входных файлов")
            wait_before_exit()
            return
        
        filepaths = sys.argv[1:]
        log(f"Получено файлов: {len(filepaths)}")
        
        # Обработка файлов
        teams_data = process_files(filepaths)
        
        if not teams_data:
            notify_error("Не удалось извлечь данные из файлов!")
            wait_before_exit()
            return
        
        total_participants = sum(len(places) for places in teams_data.values())
        log(f"Найдено команд: {len(teams_data)}, участников: {total_participants}")
        
        # Создание выходной папки
        output_folder, timestamp = get_output_folder(filepaths)
        
        # Создание xlsx
        output_filename = f"Результаты по командам {timestamp}.xlsx"
        output_path = os.path.join(output_folder, output_filename)
        
        create_xlsx(teams_data, output_path)
        
        # Уведомление об успехе
        notify("html_to_xlsx — Готово!",
               f"Команд: {len(teams_data)}, участников: {total_participants}")
        
        # Открываем папку с результатом
        open_folder(output_folder)
        
        log(f"Результаты сохранены в: {output_folder}")
        wait_before_exit()
        
    except Exception as e:
        error_msg = f"Критическая ошибка: {str(e)}"
        log(error_msg)
        log(traceback.format_exc())
        notify_error(str(e))
        wait_before_exit()


if __name__ == "__main__":
    main()